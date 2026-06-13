from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, GradientFill
from openpyxl.utils import get_column_letter
from openpyxl.styles.numbers import FORMAT_NUMBER_COMMA_SEPARATED1

wb = Workbook()

# ─── COLORS ───────────────────────────────────────────────────────────────────
BG_DARK      = "0D0D0D"
BG_PANEL     = "141414"
BG_CARD      = "1A1A2E"
ACCENT_BLUE  = "16213E"
ACCENT_TEAL  = "0F3460"
GOLD         = "E94560"
TEAL_BRIGHT  = "00D4FF"
GREEN_BRIGHT = "39FF14"
PURPLE       = "7B2FBE"
ORANGE       = "FF6B35"
YELLOW       = "FFD700"
WHITE        = "FFFFFF"
LIGHT_GRAY   = "B0B0B0"
MID_GRAY     = "888888"
DARK_GRAY    = "333333"
LEARN_GREEN  = "1B4332"
LEARN_TEXT   = "52B788"
SKIP_RED     = "3D0000"
SKIP_TEXT    = "FF6B6B"
SKIM_AMBER   = "3D2B00"
SKIM_TEXT    = "FFB347"
FREE_TEAL    = "003D3D"
FREE_TEXT    = "00D4FF"

def style_cell(ws, row, col, value="", bold=False, size=11, color=WHITE,
               bg=None, align="left", wrap=False, border=False, italic=False,
               num_fmt=None):
    cell = ws.cell(row=row, column=col, value=value)
    font_kw = {"bold": bold, "size": size, "color": color, "name": "Arial"}
    if italic:
        font_kw["italic"] = True
    cell.font = Font(**font_kw)
    if bg:
        cell.fill = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal=align, vertical="center",
                                wrap_text=wrap)
    if num_fmt:
        cell.number_format = num_fmt
    if border:
        thin = Side(border_style="thin", color="2A2A2A")
        cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
    return cell

def header_row(ws, row, values, bgs, colors, sizes, bolds, heights=None):
    for i, (val, bg, col, sz, bld) in enumerate(
            zip(values, bgs, colors, sizes, bolds), 1):
        style_cell(ws, row, i, val, bld, sz, col, bg, "center", True)
    ws.row_dimensions[row].height = heights or 28

def set_col_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

def merge_title(ws, row, col_start, col_end, value, bg, color, size, bold=True, height=36):
    ws.merge_cells(start_row=row, start_column=col_start,
                   end_row=row, end_column=col_end)
    cell = ws.cell(row=row, column=col_start, value=value)
    cell.font = Font(bold=bold, size=size, color=color, name="Arial")
    cell.fill = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[row].height = height
    return cell

def fill_row(ws, row, values_styles, height=45):
    for col, (val, bold, size, color, bg, align, wrap) in enumerate(values_styles, 1):
        style_cell(ws, row, col, val, bold, size, color, bg, align, wrap, True)
    ws.row_dimensions[row].height = height

# ══════════════════════════════════════════════════════════════════════════════
# SHEET 1 — Roadmap Overview
# ══════════════════════════════════════════════════════════════════════════════
ws1 = wb.active
ws1.title = "🗺 Roadmap Overview"
ws1.sheet_view.showGridLines = False
ws1.sheet_properties.tabColor = "00D4FF"

for i in range(1, 200):
    ws1.row_dimensions[i].height = 15

merge_title(ws1, 1, 1, 12,
    "AI ENGINEER FREE ROADMAP 2026 — YOUTUBE & OPEN SOURCE EDITION",
    BG_DARK, TEAL_BRIGHT, 18, height=48)
merge_title(ws1, 2, 1, 12,
    "Python → Job-Ready LLM/AI Engineer  •  100% Free  •  All 5 Phases  •  Best YouTube Channels + Free Courses  •  Real Projects, Zero Cost",
    BG_DARK, LIGHT_GRAY, 10, bold=False, height=28)

# Empty spacer
ws1.row_dimensions[3].height = 8

# Legend row
legend_items = [
    ("✅ LEARN", LEARN_GREEN, LEARN_TEXT),
    ("🚫 SKIP", SKIP_RED, SKIP_TEXT),
    ("⚡ SKIM", SKIM_AMBER, SKIM_TEXT),
    ("🆓 FREE", FREE_TEAL, FREE_TEXT),
]
ws1.merge_cells("A4:B4")
ws1.cell(4,1).value = "LEGEND:"
ws1.cell(4,1).font = Font(bold=True, size=9, color=MID_GRAY, name="Arial")
ws1.cell(4,1).fill = PatternFill("solid", fgColor=BG_DARK)
ws1.cell(4,1).alignment = Alignment(horizontal="right", vertical="center")
for i, (text, bg, fg) in enumerate(legend_items, 3):
    style_cell(ws1, 4, i, text, True, 9, fg, bg, "center")
ws1.row_dimensions[4].height = 20

# Header row 5
hdrs = ["Phase", "Title", "Duration\n(2h/day)", "Duration\n(4h/day)",
        "Primary YouTube Resource", "Channel / Creator", "Views /\nSubs",
        "Must Learn (Core Topics)", "Skip Entirely",
        "Project(s) to Build", "Cost", "Alt Resources"]
hdr_bgs  = [ACCENT_TEAL]*12
hdr_cols = [TEAL_BRIGHT]*12
hdr_sz   = [10]*12
hdr_bld  = [True]*12
header_row(ws1, 5, hdrs, hdr_bgs, hdr_cols, hdr_sz, hdr_bld, 36)

phase_colors = [
    ("Phase 1", "0D47A1", "1565C0"),
    ("Phase 2", "4A148C", "6A1B9A"),
    ("Phase 3", "B71C1C", "C62828"),
    ("Phase 4A","1B5E20", "2E7D32"),
    ("Phase 4B","E65100", "EF6C00"),
    ("Phase 5", "37474F", "455A64"),
]

data = [
    ("Phase 1","Python Foundations","3–4 weeks","2 weeks",
     "Python for Everybody (Full Course) — Dr. Chuck / freeCodeCamp",
     "freeCodeCamp.org","13M+ views",
     "Variables, loops, functions, OOP, file I/O, APIs, web scraping (requests/BeautifulSoup), pandas basics, error handling, virtual environments, pip packages — watch first 10 hrs only",
     "GUI with tkinter, web frameworks, database deep-dives — not needed for AI engineering",
     "Automation scripts (file renamer, email sender) | REST API caller | CSV/JSON pipeline | Web scraper with export",
     "FREE","Also: Mosh Python Tutorial (6h), CS50P (Harvard free), Corey Schafer Python Series"),

    ("Phase 2","ML Essentials (Minimum Viable)","4–5 weeks","2–3 weeks",
     "Machine Learning Course for Beginners — freeCodeCamp (scikit-learn focus)",
     "freeCodeCamp.org / Patrick Loeber","5M+ views",
     "Regression (linear, logistic), classification (decision trees, SVM), clustering (K-Means), model evaluation (confusion matrix, ROC, cross-val), scikit-learn pipeline, train/test split, overfitting concepts — Python only",
     "Deep learning sections, R sections, reinforcement learning, pre-transformer NLP (bag of words, TF-IDF), dimensionality reduction deep-dives",
     "Churn predictor with scikit-learn + ROC curve | House price regressor | K-Means customer segmentation visualised",
     "FREE","Also: StatQuest with Josh Starmer (YouTube), fast.ai Practical ML (free), Kaggle Learn (micro-courses)"),

    ("Phase 3","LLM Engineering Core — The Real Job","10–12 weeks","5–6 weeks",
     "LLM Course by Maxime Labonne (GitHub) + Andrej Karpathy 'Neural Networks: Zero to Hero'",
     "GitHub: mlabonne/llm-course + @karpathy","1M+ views / 200K+ ⭐",
     "ALL mandatory: LLM fundamentals, OpenAI/Claude/Gemini APIs, prompt engineering (system prompts, few-shot, chain-of-thought), LangChain, Hugging Face, multi-modal, RAG pipeline (chunking→embedding→vector store→retrieval→generation), QLoRA fine-tuning, Gradio UIs, AI agents with tool use, open-source models (Llama/Mistral)",
     "Nothing in this phase. Every project is portfolio gold.",
     "AI brochure generator | Multi-modal chatbot | RAG knowledge-worker | Fine-tuned 7B model | Meeting minutes generator | Price predictor (frontier vs fine-tuned)",
     "FREE","Also: DeepLearning.AI short courses (free audit), Weights & Biases courses (free), Sam Witteveen LLM YouTube series"),

    ("Phase 4A","Agentic AI Engineering","6–8 weeks","3–4 weeks",
     "LangGraph Tutorial Series — LangChain YouTube + Harrison Chase talks",
     "LangChain YouTube + Prompt Engineering","500K+ views combined",
     "Autonomous agent architecture, tool-calling design, agent memory (short-term/long-term/episodic), LangGraph state machines, multi-agent collaboration (CrewAI), MCP (Model Context Protocol), production agent deployment, failure modes & guardrails, observability with LangSmith",
     "Nothing mandatory to skip. If you already understand a concept from Phase 3, move faster.",
     "Research agent | Multi-agent content pipeline (3+ agents) | Customer support agent with memory | LangSmith-traced agent with full observability",
     "FREE","Also: CrewAI YouTube channel (free), BentoML deployment tutorials, Arize Phoenix observability (free tier)"),

    ("Phase 4B","AI Workflows & Orchestration","4–5 weeks","2 weeks",
     "n8n YouTube Channel (official) + OpenAI Agents SDK docs + tutorials",
     "n8n official YouTube + various","300K+ views",
     "n8n workflow orchestration, OpenAI Agents SDK deep dive, LangGraph advanced patterns, agent tool design, human-in-the-loop workflows, webhook triggers, multi-model routing, cost-aware agent design, connecting agents to external APIs/databases",
     "Basic LLM intro sections (you know this from Phase 3). Skim 'What is an LLM' modules.",
     "Automated content pipeline | n8n multi-step business workflow | Cost-optimised model router (big model → small model by complexity)",
     "FREE","Also: Zapier free tier (no-code alternative), Make.com tutorials, Activepieces (open-source n8n alternative)"),

    ("Phase 5","Production & Deployment (Senior Roles)","3–4 weeks","2 weeks",
     "Docker Tutorial for Beginners — TechWorld with Nana + K8s crash course",
     "TechWorld with Nana","15M+ views",
     "Docker fundamentals (images, containers, volumes), Dockerfile best practices, docker-compose for multi-service AI apps, Kubernetes core (pods, services, deployments), deploying FastAPI + LLM app to cloud (GCP Cloud Run free tier), CI/CD basics for AI systems",
     "Advanced K8s (StatefulSets, Operators, Helm deep-dives) — DevOps territory. Skip if targeting junior roles first.",
     "Dockerised RAG API (FastAPI + Chroma + OpenAI on GCP Cloud Run) | Full pipeline with health checks, env vars, secrets",
     "FREE","Also: KodeKloud free Docker/K8s courses, Play with Docker (browser labs), GCP free tier ($300 credit)"),
]

row_bgs_alt = [BG_PANEL, "111111"]
for r, (row_data, pc) in enumerate(zip(data, phase_colors), 6):
    phase, title, d2, d4, resource, channel, views, learn, skip, projects, cost, alts = row_data
    bg = row_bgs_alt[r % 2]
    phase_bg = pc[1]

    cells = [
        (phase,    True, 10, TEAL_BRIGHT, phase_bg,   "center", False),
        (title,    True, 10, WHITE,       bg,          "left",   True),
        (d2,       False,9,  LIGHT_GRAY,  bg,          "center", True),
        (d4,       False,9,  LIGHT_GRAY,  bg,          "center", True),
        (resource, True, 9,  YELLOW,      bg,          "left",   True),
        (channel,  False,9,  TEAL_BRIGHT, bg,          "left",   True),
        (views,    False,9,  GREEN_BRIGHT,bg,          "center", False),
        (learn,    False,9,  WHITE,       LEARN_GREEN, "left",   True),
        (skip,     False,9,  SKIP_TEXT,   SKIP_RED,    "left",   True),
        (projects, False,9,  ORANGE,      bg,          "left",   True),
        (cost,     True, 10, GREEN_BRIGHT,FREE_TEAL,   "center", False),
        (alts,     False,8,  MID_GRAY,    bg,          "left",   True),
    ]
    fill_row(ws1, r, cells, height=72)

set_col_widths(ws1, [10, 22, 12, 12, 30, 22, 12, 40, 35, 38, 8, 38])

# freeze panes
ws1.freeze_panes = "A6"

# ══════════════════════════════════════════════════════════════════════════════
# SHEET 2 — Course Deep Dives (YouTube)
# ══════════════════════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("📚 Resource Deep Dives")
ws2.sheet_view.showGridLines = False
ws2.sheet_properties.tabColor = "7B2FBE"

merge_title(ws2, 1, 1, 10,
    "YOUTUBE & FREE RESOURCE DEEP DIVES — EXACT TOPICS TO LEARN & SKIP",
    BG_DARK, PURPLE, 16, height=42)
merge_title(ws2, 2, 1, 10,
    "Green = Learn  |  Red = Skip  |  Orange = Skim  |  🔗 = Free Link",
    BG_DARK, LIGHT_GRAY, 9, bold=False, height=22)

hdrs2 = ["Resource / Channel","Section / Module","Status","Why","Time Est.",
         "Tools / Libraries","Concepts You'll Learn","Skip Reason","Project Output","Importance"]
header_row(ws2, 3, hdrs2,
           [ACCENT_TEAL]*10, [TEAL_BRIGHT]*10, [10]*10, [True]*10, 32)

deep_data = [
    # freeCodeCamp Python
    ("freeCodeCamp Python Full Course (Dr. Chuck)","Hrs 1–4: Python Basics","✅ LEARN","Foundation for everything","~4 hrs","Python 3, VSCode, Replit","Variables, data types, strings, lists, dicts, loops, functions, scope","—","5 mini-programs: calculator, quiz, blackjack","Critical"),
    ("freeCodeCamp Python Full Course (Dr. Chuck)","Hrs 5–10: Intermediate Python","✅ LEARN","APIs, file I/O, OOP all used daily in AI","~5 hrs","requests, pandas, BeautifulSoup, CSV/JSON","OOP classes, inheritance, file handling, web scraping, REST APIs, error handling","—","Web scraper + API project (weather CLI, news reader)","Critical"),
    ("freeCodeCamp Python Full Course (Dr. Chuck)","Hrs 11–16: Web Dev sections","🚫 SKIP","Flask/HTML/CSS not needed for AI engineering","~5 hrs","Flask, HTML, CSS","Web dev, templates, forms","You won't build web apps from scratch — use Gradio/Streamlit for AI UIs","—","None"),
    ("Corey Schafer Python Series (YouTube)","OOP Playlist (6 videos)","✅ LEARN","Deep OOP needed for LangChain & agents code","~3 hrs","Python 3","Classes, __init__, inheritance, dunder methods, properties","—","Build a reusable API client class","Critical"),

    # ML Phase 2
    ("freeCodeCamp ML Course (Patrick Loeber)","Part 1–2: Regression","✅ LEARN","Mental model for supervised learning","~2 hrs","scikit-learn, pandas, matplotlib","Linear regression, polynomial regression, model evaluation (R²), train/test split","ALL math derivation sections — watch for intuition only","House price predictor","Critical"),
    ("freeCodeCamp ML Course (Patrick Loeber)","Part 3: Classification","✅ LEARN","Core ML intuition you'll reference forever","~3 hrs","scikit-learn","Logistic regression, KNN, SVM, decision trees, random forests, confusion matrix, ROC","—","Churn predictor with ROC curve","Critical"),
    ("freeCodeCamp ML Course (Patrick Loeber)","Part 4: Clustering","✅ LEARN","Unsupervised intuition for embeddings","~2 hrs","scikit-learn, scipy","K-Means, hierarchical clustering, elbow method","—","Customer segmentation app","Important"),
    ("StatQuest with Josh Starmer","Embedding & Attention videos","✅ LEARN","Best visual explanation of transformers on YouTube","~3 hrs","none (conceptual)","How attention works, self-attention, positional encoding, why BERT/GPT architecture","Older statistics videos (SVM math) — watch only attention/transformer series for LLM path","Conceptual foundation for Phase 3","Critical"),
    ("StatQuest with Josh Starmer","Classical ML stats (SVM maths, etc.)","⚡ SKIM","Conceptually useful, not practically needed","~3 hrs","none","SVM derivations, logistic regression math","Watch only if confused about a concept from freeCodeCamp ML. Don't study systematically.","—","Low"),

    # Phase 3 LLM
    ("Andrej Karpathy — Neural Networks: Zero to Hero","All 8 videos (~14 hrs)","✅ LEARN","Best LLM intuition on the entire internet — free","~14 hrs","PyTorch (conceptual)","Backprop, language model basics, attention mechanism, GPT from scratch","Final videos on advanced GPU training if you're not a researcher","Deep understanding of why LLMs work","Critical"),
    ("mlabonne/llm-course (GitHub)","Fundamentals: LLM fundamentals section","✅ LEARN","Comprehensive map of the LLM ecosystem","~6 hrs","Various linked resources","Tokenization, context windows, temperature, prompting, model families","—","API caller comparing 3 models","Critical"),
    ("mlabonne/llm-course (GitHub)","RAG section + linked papers/notebooks","✅ LEARN","Core skill. Every AI engineer job requires RAG","~8 hrs","LangChain, ChromaDB, OpenAI/Gemini API","Chunking, embedding models, vector stores, similarity search, retrieval, reranking","—","Company knowledge-worker RAG chatbot on internal documents","Critical"),
    ("mlabonne/llm-course (GitHub)","Fine-tuning section (QLoRA notebooks)","✅ LEARN","QLoRA fine-tuning is a key CV differentiator","~8 hrs","Hugging Face PEFT, QLoRA, Weights & Biases","LoRA adapters, quantization (4-bit), training loop, evaluation, model comparison","—","Fine-tuned 7B model that beats GPT-4 on domain task","Critical"),
    ("DeepLearning.AI Short Courses (free audit)","Prompt Engineering with ChatGPT (free)","✅ LEARN","Andrew Ng + OpenAI — structured prompt patterns","~2 hrs","OpenAI API, Jupyter","System prompts, few-shot, chain-of-thought, output formatting","—","Prompt template library","Important"),
    ("DeepLearning.AI Short Courses (free audit)","LangChain for LLM Application Development","✅ LEARN","Official curriculum co-authored by Harrison Chase","~3 hrs","LangChain, OpenAI API","Chains, memory, agents, evaluation basics","LLM basics intro (you know this)","Q&A bot with document loading","Critical"),
    ("Sam Witteveen LLM YouTube Series","Gemini API + multimodal videos","✅ LEARN","Best practical Gemini API coverage on YouTube","~4 hrs","Google Gemini API, Vertex AI","Multi-modal inputs (image, audio, video), Gemini function calling, cost comparison","Older pre-Gemini-1.5 videos","Multi-modal chatbot using Gemini","Critical"),
    ("Sam Witteveen LLM YouTube Series","Older PaLM/Bard videos","🚫 SKIP","Superseded by Gemini API entirely","~3 hrs","PaLM API","Legacy API patterns","All outdated — use Gemini API instead","—","None"),

    # Phase 4A Agents
    ("LangChain YouTube — LangGraph Tutorials","LangGraph Quickstart + State Machine videos","✅ LEARN","Official source — most up-to-date LangGraph content","~5 hrs","LangGraph, LangChain","Nodes, edges, state schema, conditional routing, agent loop design, ReAct pattern","—","Single-tool research agent","Critical"),
    ("LangChain YouTube — LangGraph Tutorials","Memory systems video series","✅ LEARN","Memory is what makes agents useful vs stateless chatbots","~3 hrs","LangGraph, vector stores","Short-term (conversation) memory, long-term (vector) memory, episodic memory","—","Agent with persistent memory across sessions","Critical"),
    ("CrewAI YouTube Channel","Multi-Agent pipeline tutorials","✅ LEARN","Multi-agent is the dominant 2026 production pattern","~3 hrs","CrewAI, LangGraph","Orchestrator-worker pattern, agent roles, inter-agent comms, state passing, error recovery","Getting started intro (skim quickly)","3-agent pipeline: planner → executor → critic","Critical"),
    ("Prompt Engineering YouTube / Dave Ebbelaar","MCP Server implementation tutorials","✅ LEARN","MCP is what interviewers ask about in 2026","~2 hrs","MCP SDK, Claude API","MCP server/client architecture, tool exposure via MCP, connecting agents to external systems","—","Custom MCP server exposing internal tools","Critical"),
    ("Arize Phoenix YouTube / LangSmith docs","Observability and tracing videos","✅ LEARN","Tracing separates juniors from hireables","~2 hrs","LangSmith (free tier), Arize Phoenix (open-source)","Tracing agent loops, debugging tool calls, cost tracking per run, latency measurement","—","Fully traced agent with LangSmith dashboard","Critical"),

    # Phase 4B Workflows
    ("n8n Official YouTube Channel","Workflow fundamentals + AI node tutorials","✅ LEARN","n8n used in production companies; open-source","~4 hrs","n8n (self-hosted, free)","Trigger nodes, action nodes, webhook setup, conditional logic, AI integration","—","Email → summarise → reply automation workflow","Important"),
    ("n8n Official YouTube Channel","Advanced AI agent workflows","✅ LEARN","Connects n8n to LLMs for business automation","~3 hrs","n8n + OpenAI API","Multi-step AI pipelines, human-in-the-loop, error handling","Intro 'what is n8n' videos (skim)","Multi-step business intelligence pipeline","Critical"),
    ("Various — OpenAI Agents SDK docs & videos","OpenAI Agents SDK tutorials","✅ LEARN","Official SDK; companies on OpenAI use this","~4 hrs","openai-agents SDK","Agent traces, handoffs, guardrails, tool registration, streaming output","Skim any LLM basics intro (redundant)","Multi-step research agent using OpenAI SDK","Critical"),

    # Phase 5 Docker/K8s
    ("TechWorld with Nana — Docker Tutorial","Docker Fundamentals (full video)","✅ LEARN","Best free Docker content — clear, visual, updated","~4 hrs","Docker Desktop, Docker Hub","Images vs containers, Dockerfile syntax, layers, volumes, env vars, docker-compose","—","Dockerised FastAPI + LLM app","Critical"),
    ("TechWorld with Nana — K8s Tutorial","Kubernetes Crash Course","✅ LEARN","K8s expected for senior roles","~4 hrs","kubectl, minikube, GKE","Pods, services, deployments, ConfigMaps, Secrets, ingress, autoscaling","—","RAG API on GCP GKE with autoscaling","Important"),
    ("TechWorld with Nana — K8s Tutorial","Advanced K8s (Helm, Operators)","🚫 SKIP","DevOps territory — not AI engineering","~4 hrs","Helm, Operators","Advanced networking, CRDs","Unless targeting MLOps/DevOps hybrid role","—","Low"),
]

status_styles = {
    "✅ LEARN": (LEARN_GREEN, LEARN_TEXT),
    "🚫 SKIP":  (SKIP_RED,   SKIP_TEXT),
    "⚡ SKIM":  (SKIM_AMBER, SKIM_TEXT),
}
importance_colors = {
    "Critical":  GREEN_BRIGHT,
    "Important": YELLOW,
    "Low":       SKIP_TEXT,
    "None":      MID_GRAY,
}

prev_resource = None
for r, row in enumerate(deep_data, 4):
    res, section, status, why, time_est, tools, concepts, skip_reason, output, importance = row
    sbg, sfg = status_styles.get(status, (BG_PANEL, WHITE))
    bg = BG_PANEL if r % 2 == 0 else "111111"
    imp_color = importance_colors.get(importance, WHITE)
    res_color = TEAL_BRIGHT if res != prev_resource else LIGHT_GRAY
    prev_resource = res

    cells = [
        (res,         True, 9,  res_color, bg,        "left",  True),
        (section,     True, 9,  WHITE,     bg,        "left",  True),
        (status,      True, 9,  sfg,       sbg,       "center",False),
        (why,         False,8,  LIGHT_GRAY,bg,        "left",  True),
        (time_est,    False,9,  ORANGE,    bg,        "center",False),
        (tools,       False,8,  TEAL_BRIGHT,bg,       "left",  True),
        (concepts,    False,8,  WHITE,     bg,        "left",  True),
        (skip_reason, False,8,  SKIP_TEXT, SKIP_RED if skip_reason != "—" else bg, "left", True),
        (output,      False,8,  ORANGE,    bg,        "left",  True),
        (importance,  True, 9,  imp_color, bg,        "center",False),
    ]
    fill_row(ws2, r, cells, height=60)

set_col_widths(ws2, [28, 30, 10, 22, 8, 25, 38, 30, 30, 10])
ws2.freeze_panes = "A4"

# ══════════════════════════════════════════════════════════════════════════════
# SHEET 3 — Timeline & Cost
# ══════════════════════════════════════════════════════════════════════════════
ws3 = wb.create_sheet("⏱ Timeline & Cost")
ws3.sheet_view.showGridLines = False
ws3.sheet_properties.tabColor = "FFD700"

merge_title(ws3, 1, 1, 9,
    "TIMELINE & COST BREAKDOWN — FREE PATH (YOUTUBE + OPEN SOURCE)",
    BG_DARK, YELLOW, 16, height=42)
merge_title(ws3, 2, 1, 9,
    "Total course cost: ₹0 • You only pay for API usage (keep it minimal with smart model choice)",
    BG_DARK, GREEN_BRIGHT, 10, bold=False, height=24)

ws3.row_dimensions[3].height = 10

# Timeline section
merge_title(ws3, 4, 1, 9, "LEARNING TIMELINE BY PACE", ACCENT_TEAL, TEAL_BRIGHT, 12, height=28)
header_row(ws3, 5,
    ["Phase","Description","Hrs of Video","Hrs of Practice","Total Hours","2h/day (weeks)","4h/day (weeks)","Cumulative 2h/day","Cumulative 4h/day"],
    [ACCENT_TEAL]*9, [TEAL_BRIGHT]*9, [10]*9, [True]*9, 30)

timeline = [
    ("Phase 1","Python Foundations (freeCodeCamp + Corey Schafer)","15","20","35","3–4 wks","2 wks","3–4 wks","2 wks"),
    ("Phase 2","ML Essentials (freeCodeCamp + StatQuest)","12","15","27","3–4 wks","2 wks","6–8 wks","4 wks"),
    ("Phase 3","LLM Engineering Core (mlabonne + Karpathy + DeepLearning.AI)","30","45","75","10–12 wks","5–6 wks","16–20 wks","9–10 wks"),
    ("Phase 4A","Agentic AI (LangChain YT + CrewAI + MCP tutorials)","14","25","39","5–7 wks","3–4 wks","21–27 wks","12–14 wks"),
    ("Phase 4B","AI Workflows (n8n YT + OpenAI Agents SDK)","14","20","34","4–5 wks","2 wks","25–32 wks","14–16 wks"),
    ("Phase 5","Docker & K8s (TechWorld with Nana)","10","15","25","3–4 wks","2 wks","28–36 wks","16–18 wks"),
    ("TOTAL","Phases 1–5 complete","95","140","235","~28–36 wks","~16–18 wks","~8 months","~5 months"),
]
phase_bgs_tl = [
    "0D47A1","4A148C","B71C1C","1B5E20","E65100","37474F",ACCENT_TEAL
]
for i, (row, pbg) in enumerate(zip(timeline, phase_bgs_tl), 6):
    bg = BG_PANEL if i % 2 == 0 else "111111"
    is_total = row[0] == "TOTAL"
    for col, val in enumerate(row, 1):
        c_color = TEAL_BRIGHT if col == 1 else (GREEN_BRIGHT if is_total else WHITE)
        c_bg = pbg if col == 1 else (ACCENT_TEAL if is_total else bg)
        style_cell(ws3, i, col, val, is_total or col==1, 10, c_color, c_bg, "center", True, True)
    ws3.row_dimensions[i].height = 28

ws3.row_dimensions[13].height = 16

# Free Resources section
merge_title(ws3, 14, 1, 9, "FREE RESOURCE LINKS BY PHASE", ACCENT_TEAL, TEAL_BRIGHT, 12, height=28)
header_row(ws3, 15,
    ["Phase","Resource Name","Creator / Source","URL / Where to Find","Hours","Phase","Type","Free?","Notes"],
    [ACCENT_TEAL]*9, [TEAL_BRIGHT]*9, [10]*9, [True]*9, 28)

free_resources = [
    ("Phase 1","Python for Everybody Full Course","Dr. Chuck / freeCodeCamp","youtube.com/watch?v=8DvywoWv6fI","16h","1","Full course","🆓 100% Free","Best for total beginners"),
    ("Phase 1","Python OOP Series","Corey Schafer / YouTube","youtube.com — search 'Corey Schafer Python OOP'","3h","1","Series","🆓 100% Free","Deep OOP coverage"),
    ("Phase 1","CS50P (Intro to Python)","Harvard / edX free audit","cs50.harvard.edu/python","10h","1","Full course","🆓 Free audit","Rigorous Harvard quality"),
    ("Phase 2","ML Course for Beginners","Patrick Loeber / freeCodeCamp","youtube.com/watch?v=NWONeJKn6kc","6h","2","Full course","🆓 100% Free","Best free ML intro"),
    ("Phase 2","StatQuest ML/Stats Series","Josh Starmer / YouTube","statquest.org","20h","2","Series","🆓 100% Free","Best conceptual ML explanations"),
    ("Phase 2","Kaggle Learn — ML Micro-courses","Kaggle","kaggle.com/learn","8h","2","Micro-courses","🆓 100% Free","Hands-on with datasets"),
    ("Phase 3","Neural Networks: Zero to Hero","Andrej Karpathy / YouTube","karpathy.ai / YouTube","14h","3","Deep course","🆓 100% Free","Best LLM intuition on the internet"),
    ("Phase 3","LLM Course (GitHub)","Maxime Labonne / mlabonne","github.com/mlabonne/llm-course","30h+","3","Curated roadmap","🆓 100% Free","Comprehensive LLM map with notebooks"),
    ("Phase 3","Short Courses (free audit)","DeepLearning.AI / Andrew Ng","deeplearning.ai/short-courses","15h","3","Structured courses","🆓 Free audit","Authoritative, co-created with OpenAI/Google"),
    ("Phase 3","LLM YouTube Series","Sam Witteveen","youtube.com — search 'Sam Witteveen LLM'","10h","3","Series","🆓 100% Free","Best Gemini + practical API coverage"),
    ("Phase 4A","LangGraph Tutorial Series","LangChain YouTube","youtube.com/c/LangChain","6h","4A","Official tutorials","🆓 100% Free","Always up-to-date with latest LangGraph"),
    ("Phase 4A","CrewAI Tutorials","CrewAI YouTube","youtube.com/crewai","4h","4A","Official tutorials","🆓 100% Free","Best multi-agent framework tutorials"),
    ("Phase 4A","LangSmith Observability","LangSmith docs + YouTube","docs.smith.langchain.com","2h","4A","Docs + video","🆓 Free tier","Essential for production-grade agents"),
    ("Phase 4B","n8n Automation Tutorials","n8n Official YouTube","youtube.com/n8n","8h","4B","Official tutorials","🆓 100% Free","Self-hosted n8n = zero cost"),
    ("Phase 4B","OpenAI Agents SDK Docs","OpenAI","platform.openai.com/docs/guides/agents","3h","4B","Official docs","🆓 100% Free","Read the docs, build the examples"),
    ("Phase 5","Docker Tutorial for Beginners","TechWorld with Nana","youtube.com — search 'Nana Docker Tutorial'","4h","5","Full tutorial","🆓 100% Free","Best free Docker content on YouTube"),
    ("Phase 5","Kubernetes Crash Course","TechWorld with Nana","youtube.com — search 'Nana Kubernetes Crash Course'","4h","5","Full tutorial","🆓 100% Free","Same quality as paid Udemy courses"),
    ("Phase 5","Play with Docker","Docker Labs","labs.play-with-docker.com","∞","5","Browser lab","🆓 100% Free","Practice Docker without local install"),
]

for i, row in enumerate(free_resources, 16):
    bg = BG_PANEL if i % 2 == 0 else "111111"
    phase, name, creator, url, hours, ph, rtype, free, notes = row
    cells_data = [
        (phase,   True, 9,  TEAL_BRIGHT, bg, "center", False),
        (name,    True, 9,  WHITE,       bg, "left",   True),
        (creator, False,9,  TEAL_BRIGHT, bg, "left",   True),
        (url,     False,8,  YELLOW,      bg, "left",   True),
        (hours,   False,9,  ORANGE,      bg, "center", False),
        (ph,      False,9,  LIGHT_GRAY,  bg, "center", False),
        (rtype,   False,8,  LIGHT_GRAY,  bg, "center", True),
        (free,    True, 9,  GREEN_BRIGHT,FREE_TEAL,"center",False),
        (notes,   False,8,  MID_GRAY,    bg, "left",   True),
    ]
    for col, (val, bold, size, color, cbg, align, wrap) in enumerate(cells_data, 1):
        style_cell(ws3, i, col, val, bold, size, color, cbg, align, wrap, True)
    ws3.row_dimensions[i].height = 30

ws3.row_dimensions[35].height = 16

# API Cost section
merge_title(ws3, 36, 1, 9, "API COSTS TO BUDGET (This is your only real expense)", ACCENT_TEAL, YELLOW, 12, height=28)
header_row(ws3, 37,
    ["Expense","Provider","Monthly Cost","When Needed","Free Tier?","Tips to Minimise","Total Budget","Priority","Notes"],
    [ACCENT_TEAL]*9, [TEAL_BRIGHT]*9, [10]*9, [True]*9, 28)

api_costs = [
    ("OpenAI API Credits","OpenAI","$5–20/mo","Phase 3 onwards","No ($5 free trial)","Use gpt-4o-mini for dev (10× cheaper); GPT-4o only for final tests. Set hard limits.","~$30–80 over 6 months","Essential","gpt-4o-mini is good enough for 80% of dev work — don't use GPT-4o until needed"),
    ("Anthropic API Credits","Anthropic","$2–10/mo","Phase 3 onwards","No ($5 free trial)","Claude Haiku is 10× cheaper than Sonnet. Use Haiku for dev; Sonnet for final demos.","~$15–40 over 6 months","Important","Only need when comparing models or building Claude-specific features (MCP)"),
    ("Google Colab (free tier)","Google","$0","Phase 3 fine-tuning","Yes — but limited GPU","Free tier gives T4 GPU (~30 min sessions). Enough for small QLoRA runs if patient.","$0","Free","If free tier too slow, Colab Pro ($10/mo) for 2–3 months only = ~$20 total"),
    ("Hugging Face","Hugging Face","$0","Phase 3–4A","Yes — generous free tier","Free inference API + free model hosting. Zero cost for learning.","$0","Free","HF Spaces lets you deploy Gradio demos free — great for portfolio"),
    ("ChromaDB / Qdrant","Chroma / Qdrant","$0","Phase 3 RAG","Yes — local install","Run ChromaDB locally (pip install). Qdrant has a free cloud tier (1GB).","$0","Free","No reason to pay for a vector DB during learning phase"),
    ("n8n","n8n.io","$0","Phase 4B","Yes — self-hosted","Self-host n8n on your local machine or on a free Railway/Render tier.","$0","Free","n8n Cloud has paid plans but self-hosted is always free"),
    ("LangSmith","LangChain","$0","Phase 4A","Yes — free tier","Free tier: 5,000 traces/month. More than enough for learning.","$0","Free","Upgrade only if deploying a production app with heavy usage"),
    ("Vercel / Railway deployment","Vercel / Railway","$0","Phase 5","Yes — free tier","Deploy demos on Vercel/Railway free tier. A live portfolio URL = 10× better than GitHub.","$0 – $5","Optional","Only needed if you want a live portfolio URL (worth it for job applications)"),
    ("GCP Free Trial","Google Cloud","$0 (then $10–30)","Phase 5","Yes — $300 credit","$300 free credit lasts 6–12 months for learning usage. Cloud Run free tier after.","$0 (then $0–30)","Optional","$300 credit handles all Phase 5 deployment learning with budget to spare"),
]

for i, row in enumerate(api_costs, 38):
    bg = BG_PANEL if i % 2 == 0 else "111111"
    for col, (val, bold, size, color, cbg, align, wrap) in enumerate([
        (row[0], True,  9,  WHITE,        bg,        "left",   True),
        (row[1], False, 9,  TEAL_BRIGHT,  bg,        "left",   False),
        (row[2], True,  9,  ORANGE,       bg,        "center", False),
        (row[3], False, 8,  LIGHT_GRAY,   bg,        "left",   True),
        (row[4], True,  9,  GREEN_BRIGHT, FREE_TEAL, "center", True),
        (row[5], False, 8,  LIGHT_GRAY,   bg,        "left",   True),
        (row[6], True,  9,  GREEN_BRIGHT, bg,        "center", True),
        (row[7], True,  9,  YELLOW,       bg,        "center", False),
        (row[8], False, 8,  MID_GRAY,     bg,        "left",   True),
    ], 1):
        style_cell(ws3, i, col, val, bold, size, color, cbg, align, wrap, True)
    ws3.row_dimensions[i].height = 40

set_col_widths(ws3, [28, 28, 14, 22, 14, 35, 16, 12, 40])
ws3.freeze_panes = "A6"

# ══════════════════════════════════════════════════════════════════════════════
# SHEET 4 — Udemy vs Free Comparison
# ══════════════════════════════════════════════════════════════════════════════
ws4 = wb.create_sheet("⚖ Paid vs Free")
ws4.sheet_view.showGridLines = False
ws4.sheet_properties.tabColor = "E94560"

merge_title(ws4, 1, 1, 8,
    "PAID (UDEMY) vs FREE (YOUTUBE) — FULL COMPARISON PER PHASE",
    BG_DARK, GOLD, 16, height=42)
merge_title(ws4, 2, 1, 8,
    "Verdict: Free path covers 90% of the same content. Paid path wins on structure and project guidance. Both paths lead to the same job.",
    BG_DARK, LIGHT_GRAY, 9, bold=False, height=24)

header_row(ws4, 3,
    ["Phase / Topic","Paid Option (Udemy)","Free Option (YouTube / GitHub)","Content Parity","Where Free Wins","Where Paid Wins","Verdict","Recommendation"],
    [ACCENT_TEAL]*8, [TEAL_BRIGHT]*8, [10]*8, [True]*8, 32)

compare_data = [
    ("Phase 1 — Python",
     "Angela Yu 100 Days (₹499)\nStructured 100-day course, project every day, polished production",
     "freeCodeCamp Python (Dr. Chuck) + Corey Schafer OOP series + CS50P (Harvard)\nTotal: ~30h free content",
     "95%","Free path has same topics. CS50P is more rigorous than Angela Yu for fundamentals.",
     "Angela Yu's project-per-day structure forces daily coding habit. Harder to replicate with YouTube.",
     "🟡 TIE (lean free)","If you're self-disciplined: go free. If you need structure to stay consistent: buy Angela Yu on sale."),

    ("Phase 2 — ML Essentials",
     "ML A-Z (₹499)\nKirill Eremenko — comprehensive, both Python + R",
     "freeCodeCamp ML (Patrick Loeber) + StatQuest + Kaggle Learn\nTotal: ~25h free content",
     "90%","StatQuest explanations of ML concepts are clearer than ML A-Z. Kaggle gives hands-on practice.",
     "ML A-Z has one cohesive structure. Free path requires you to jump between sources.",
     "🟢 FREE WINS","Go free. StatQuest + Patrick Loeber + Kaggle is objectively better than ML A-Z for LLM prep."),

    ("Phase 3 — LLM Engineering Core",
     "Ed Donner LLM Engineering (₹499)\n8 weeks, 8 projects, frontier APIs, RAG, QLoRA fine-tuning",
     "mlabonne/llm-course (GitHub) + Karpathy Zero to Hero + DeepLearning.AI short courses\nTotal: ~50h free content",
     "85%","Karpathy's Zero to Hero is better for LLM fundamentals than any paid course. mlabonne roadmap is comprehensive.",
     "Ed Donner's 8 structured projects are a complete portfolio out of the box. Hard to replicate that cohesion for free.",
     "🔴 PAID WINS (but barely)","Ed Donner's course is the one paid course most worth buying. If you can't afford it, the free path still works — but you'll need more self-direction on projects."),

    ("Phase 4A — Agentic AI",
     "Ed Donner Agentic AI Course (₹499)\nDirect continuation of Phase 3 course, MCP deep-dive",
     "LangChain YouTube + CrewAI channel + LangSmith docs + MCP tutorials\nTotal: ~20h free content",
     "80%","LangChain official YouTube is always more up-to-date than any recorded course. MCP tutorials on YouTube are current.",
     "Ed Donner covers MCP + production deployment in one cohesive arc. Free content is fragmented.",
     "🟡 TIE (lean free)","If you bought Ed Donner Phase 3, buy Phase 4A for continuity. Otherwise, the free path with LangChain YouTube + CrewAI channel is excellent."),

    ("Phase 4B — AI Workflows",
     "Schwarzmüller AI Agents & Workflows (₹499)\nAcademind quality, n8n + OpenAI SDK",
     "n8n Official YouTube + OpenAI Agents SDK docs + various tutorials\nTotal: ~15h free content",
     "85%","n8n's official YouTube covers their own tool better than any third-party course. OpenAI docs are authoritative.",
     "Schwarzmüller course covers n8n + LangGraph + OpenAI SDK in one place with consistent style.",
     "🟢 FREE WINS","Go free. n8n has excellent official YouTube content. OpenAI SDK docs are well-written. No need to pay."),

    ("Phase 5 — Docker & K8s",
     "Stephen Grider Docker & K8s (₹499)\nPolished, structured, cloud deployment focus",
     "TechWorld with Nana Docker + K8s tutorials + KodeKloud free + Play with Docker\nTotal: ~15h free content",
     "95%","TechWorld with Nana is arguably better than Grider. KodeKloud labs are hands-on and free.",
     "Grider's course is more polished with explicit AI system deployment examples.",
     "🟢 FREE WINS","TechWorld with Nana is excellent. Skip Grider. Use the saved ₹499 for API credits."),

    ("OVERALL VERDICT",
     "Total: ₹2,495–₹2,994 (~$50–60 USD)\nMost value-dense paid path available",
     "Total: ₹0 (~$0)\nSame career outcome, less structure",
     "88%","Free path saves ₹2,994. That money covers 3+ months of API usage.",
     "Paid path (especially Ed Donner courses) gives structured projects that build a portfolio efficiently.",
     "🟡 PERSONAL CHOICE","TIGHT BUDGET: Go 100% free — the outcome is the same if you're self-disciplined. SOME BUDGET: Buy Ed Donner Phase 3 + 4A only (~₹998). FULL BUDGET: All 6 Udemy courses for structure + portfolio."),
]

parity_colors = {"95%": GREEN_BRIGHT, "90%": GREEN_BRIGHT, "88%": GREEN_BRIGHT,
                 "85%": YELLOW, "80%": ORANGE}

for i, row in enumerate(compare_data, 4):
    phase, paid, free, parity, free_win, paid_win, verdict, rec = row
    bg = BG_PANEL if i % 2 == 0 else "111111"
    is_total = phase.startswith("OVERALL")
    par_color = parity_colors.get(parity, LIGHT_GRAY)

    verdict_bg = {"🟢 FREE WINS": FREE_TEAL, "🔴 PAID WINS (but barely)": SKIP_RED,
                  "🟡 TIE (lean free)": SKIM_AMBER, "🟡 PERSONAL CHOICE": SKIM_AMBER}.get(verdict, bg)
    verdict_fg = {"🟢 FREE WINS": GREEN_BRIGHT, "🔴 PAID WINS (but barely)": SKIP_TEXT,
                  "🟡 TIE (lean free)": SKIM_TEXT, "🟡 PERSONAL CHOICE": SKIM_TEXT}.get(verdict, WHITE)

    cells = [
        (phase,    True,  10, TEAL_BRIGHT, ACCENT_TEAL if is_total else ACCENT_BLUE, "left",   True),
        (paid,     False, 8,  LIGHT_GRAY,  bg,           "left",   True),
        (free,     False, 8,  GREEN_BRIGHT,bg,            "left",   True),
        (parity,   True,  10, par_color,   bg,            "center", False),
        (free_win, False, 8,  GREEN_BRIGHT,bg,            "left",   True),
        (paid_win, False, 8,  ORANGE,      bg,            "left",   True),
        (verdict,  True,  9,  verdict_fg,  verdict_bg,    "center", True),
        (rec,      False, 8,  WHITE,       bg,            "left",   True),
    ]
    fill_row(ws4, i, cells, height=70)

set_col_widths(ws4, [20, 30, 30, 10, 30, 28, 18, 35])
ws4.freeze_panes = "A4"

# ══════════════════════════════════════════════════════════════════════════════
# SHEET 5 — Interview Skills (same as Udemy sheet but with free study resources)
# ══════════════════════════════════════════════════════════════════════════════
ws5 = wb.create_sheet("🎯 Interview Skills")
ws5.sheet_view.showGridLines = False
ws5.sheet_properties.tabColor = "39FF14"

merge_title(ws5, 1, 1, 9,
    "AI ENGINEER INTERVIEW SKILLS 2026 — FREE STUDY RESOURCES MAPPED",
    BG_DARK, GREEN_BRIGHT, 16, height=42)
merge_title(ws5, 2, 1, 9,
    "You will NOT be asked to derive softmax. You WILL be asked to design a RAG system, debug an agent loop, and estimate cost per conversation.",
    BG_DARK, LIGHT_GRAY, 9, bold=False, height=24)

header_row(ws5, 3,
    ["Topic Cluster","Specific Skill","Frequency","Sample Interview Question","Expected Answer",
     "Free Resource to Study","Where Beginners Fail","Phase"],
    [ACCENT_TEAL]*8, [TEAL_BRIGHT]*8, [10]*8, [True]*8, 32)

interview_data = [
    ("RAG Architecture","Chunking strategy tradeoffs","⭐⭐⭐⭐⭐",
     "How would you chunk a 200-page PDF for a legal Q&A system?",
     "Design answer: justify chunk size by query type, overlap for context continuity, semantic vs fixed-size chunking",
     "LangChain docs on text splitters (free) | RAGAS paper (free PDF) | mlabonne RAG notebook (GitHub free)",
     "Memorise '512 tokens' without explaining why — shows no real experience","Phase 3"),
    ("RAG Architecture","Dense vs sparse retrieval","⭐⭐⭐⭐",
     "When would you use BM25 over vector search, and when both?",
     "Hybrid search: BM25 for keyword precision, vectors for semantic; hybrid (RRF fusion) for production",
     "Pinecone blog: Hybrid Search (free) | BM25 Wikipedia | Weaviate hybrid search docs (free)",
     "Only know vector search; can't explain BM25 or hybrid retrieval","Phase 3"),
    ("RAG Architecture","Reranking","⭐⭐⭐",
     "Explain the role of a cross-encoder reranker and the latency tradeoff",
     "Cross-encoder scores query-doc pairs (more accurate but slow); use after retrieval to rerank top-K",
     "Cohere Rerank blog post (free) | sentence-transformers docs (free) | BEIR benchmark paper",
     "Never implemented reranking; think retrieval quality = embedding quality alone","Phase 3"),
    ("RAG Architecture","RAGAS evaluation metrics","⭐⭐⭐⭐",
     "How do you measure whether your RAG system is hallucinating?",
     "Faithfulness (answer vs docs), Answer Relevance, Context Precision — all measured by RAGAS",
     "ragas.io documentation (free) | RAGAS GitHub (free) | LangSmith eval guide (free)",
     "Build demos, never measure them. Can't name a single metric.","Phase 4A"),
    ("RAG Architecture","Hallucination mitigation","⭐⭐⭐⭐",
     "Your RAG system is making up citations. Walk me through 3 things you'd try.",
     "Attribution prompting, citation validation post-gen, faithfulness filter, tighter retrieval, context format",
     "Anthropic prompt engineering guide (free) | LlamaIndex hallucination cookbook (GitHub free)",
     "Only say 'improve the prompt' — no systematic approach","Phase 3"),
    ("Agentic Systems","Agent vs chain distinction","⭐⭐⭐⭐⭐",
     "What makes a system truly agentic vs a deterministic LLM chain?",
     "Agent decides next step based on output; chain is fixed. Key: dynamic tool selection, self-reflection",
     "LangGraph docs (free) | Harrison Chase talks on YouTube (free) | LangChain conceptual guide",
     "Call everything an 'agent'; can't articulate the control flow difference","Phase 4A"),
    ("Agentic Systems","Tool-calling design","⭐⭐⭐⭐",
     "Design the tool schema for an agent that can search, query SQL, and send emails",
     "Tool name, description (key!), parameters (typed JSON schema), return type, error handling",
     "OpenAI function calling docs (free) | Anthropic tool use docs (free) | LangChain tools guide",
     "Write tools without descriptions; LLMs pick wrong tools","Phase 4A"),
    ("Agentic Systems","Agent failure modes","⭐⭐⭐⭐",
     "Your agent is stuck in an infinite tool-calling loop. Debug and prevent this.",
     "Max iterations limit, loop detection, LangSmith trace, tool call history inspection, stopping conditions",
     "LangGraph error handling docs (free) | LangSmith tracing tutorial (YouTube, free) | GitHub issues",
     "Never shipped an agent; can't describe a real failure they've debugged","Phase 4A"),
    ("Agentic Systems","Multi-agent design","⭐⭐⭐",
     "Design a 3-agent content pipeline from scratch",
     "Orchestrator → worker pattern, state schema, inter-agent handoffs, error recovery, human-in-the-loop",
     "CrewAI docs (free) | CrewAI YouTube channel (free) | LangGraph multi-agent tutorial",
     "Describe agents as 'just more prompts' — no architecture thinking","Phase 4A"),
    ("Prompt Engineering","System prompt design","⭐⭐⭐⭐",
     "Write a system prompt for a customer support agent that should never reveal pricing",
     "Role definition, constraints, tone, fallback behaviour, forbidden topic handling",
     "Anthropic prompt engineering guide (free) | OpenAI system prompt best practices (free)",
     "Vague prompts like 'You are a helpful assistant'","Phase 3"),
    ("Prompt Engineering","Prompt injection defense","⭐⭐⭐",
     "A user submits a doc with 'Ignore all instructions'. How do you handle it?",
     "Input sanitisation, instruction hierarchy (system > user), output validation, sandboxing",
     "OWASP LLM Top 10 (free PDF) | Simon Willison blog on prompt injection (free)",
     "Never thought about adversarial inputs; no security mindset","Phase 3–4"),
    ("Prompt Engineering","Structured output","⭐⭐⭐",
     "How do you reliably get JSON output from an LLM in production?",
     "JSON mode / response_format, Pydantic schema validation, retry on parse failure, few-shot JSON",
     "OpenAI structured outputs docs (free) | Pydantic docs (free) | Instructor library (GitHub, free)",
     "Use regex on LLM output; breaks constantly in production","Phase 3"),
    ("Evals & Observability","LLM-as-judge","⭐⭐⭐⭐⭐",
     "Build an LLM-as-judge eval for your RAG system",
     "Golden set creation, judge prompt design, scoring rubric, bias mitigation, calibration",
     "LangSmith eval docs (free) | RAGAS docs (free) | Hamel Husain eval blog posts (free)",
     "Never built an eval; say they 'tested manually'","Phase 4A"),
    ("Evals & Observability","LangSmith tracing","⭐⭐⭐⭐",
     "What does a LangSmith trace tell you that a print statement doesn't?",
     "Full DAG of calls, token usage per step, latency breakdown, tool inputs/outputs, cost per run",
     "LangSmith docs (free) | LangChain YouTube LangSmith tutorial (free) | free tier: 5k traces/mo",
     "Never used observability tooling; can't debug distributed LLM systems","Phase 4A"),
    ("System Design","RAG at scale","⭐⭐⭐⭐⭐",
     "Design a RAG assistant for 500K docs, <1.5s p95 latency, strict PII policy",
     "Chunking + metadata filtering, HNSW index, hybrid search, PII scrubber, async retrieval",
     "pgvector GitHub (free) | Qdrant docs (free) | Weaviate architecture blog (free)",
     "Design a toy system; ignore latency, cost, security entirely","Phase 3–4"),
    ("System Design","Cost modeling","⭐⭐⭐⭐",
     "Estimate monthly API cost of a 10-turn chat app with 10,000 daily users",
     "(input_tokens + output_tokens) × price × turns × users × days; identify caching opportunities",
     "OpenAI pricing page (free) | Anthropic pricing page (free) | tokencounter.org (free tool)",
     "Never calculated cost; bill shock is a production disaster","Phase 3"),
    ("Fine-Tuning","When to fine-tune","⭐⭐⭐⭐",
     "A client wants their chatbot to always respond in formal Hindi. Fine-tune or prompt?",
     "Prompt first; fine-tune only if prompt can't achieve consistency after 100+ examples OR cost/latency prohibitive",
     "Anthropic fine-tuning docs (free) | OpenAI fine-tuning guide (free) | Hugging Face PEFT docs",
     "Immediately suggest fine-tuning; don't know the cost/time tradeoff","Phase 3"),
    ("Fine-Tuning","QLoRA mechanics","⭐⭐⭐",
     "What is QLoRA and why does it make fine-tuning accessible?",
     "Quantize base model to 4-bit (reduces memory 4×); add LoRA adapters; train only adapters; merge back",
     "QLoRA paper on arXiv (free) | Hugging Face PEFT docs (free) | Sebastian Raschka QLoRA blog (free)",
     "Know the name but can't explain the mechanics","Phase 3"),
    ("MCP / Tooling","MCP architecture","⭐⭐⭐",
     "What is the Model Context Protocol and when use it vs function calling?",
     "MCP = open standard for AI-to-tool communication; use when multiple AI systems share tools",
     "MCP official specification (free) | Anthropic MCP blog (free) | MCP GitHub SDK examples",
     "Never heard of MCP; it's a 2025 addition many candidates miss","Phase 4A"),
]

cluster_colors = {
    "RAG Architecture": "1B4332",
    "Agentic Systems": "1B2A4A",
    "Prompt Engineering": "3D2B00",
    "Evals & Observability": "2D0A4E",
    "System Design": "3D1A00",
    "Fine-Tuning": "1A3D00",
    "MCP / Tooling": "003D3D",
}
cluster_text = {
    "RAG Architecture": LEARN_TEXT,
    "Agentic Systems": TEAL_BRIGHT,
    "Prompt Engineering": SKIM_TEXT,
    "Evals & Observability": PURPLE,
    "System Design": ORANGE,
    "Fine-Tuning": GREEN_BRIGHT,
    "MCP / Tooling": FREE_TEXT,
}

for i, row in enumerate(interview_data, 4):
    cluster, skill, freq, q, answer, free_res, fail, phase = row
    bg = BG_PANEL if i % 2 == 0 else "111111"
    cbg = cluster_colors.get(cluster, BG_PANEL)
    cfg = cluster_text.get(cluster, WHITE)

    cells = [
        (cluster,  True,  9,  cfg,         cbg, "center", True),
        (skill,    True,  9,  WHITE,        bg,  "left",   True),
        (freq,     True,  10, YELLOW,       bg,  "center", False),
        (q,        False, 8,  LIGHT_GRAY,   bg,  "left",   True),
        (answer,   False, 8,  WHITE,        bg,  "left",   True),
        (free_res, False, 8,  GREEN_BRIGHT, FREE_TEAL, "left", True),
        (fail,     False, 8,  SKIP_TEXT,    SKIP_RED, "left", True),
        (phase,    True,  9,  TEAL_BRIGHT,  ACCENT_TEAL, "center", False),
    ]
    fill_row(ws5, i, cells, height=60)

set_col_widths(ws5, [18, 22, 10, 35, 38, 38, 32, 10])
ws5.freeze_panes = "A4"

# ══════════════════════════════════════════════════════════════════════════════
# SHEET 6 — Top Free Resources Ranked
# ══════════════════════════════════════════════════════════════════════════════
ws6 = wb.create_sheet("🏆 Top Free Resources")
ws6.sheet_view.showGridLines = False
ws6.sheet_properties.tabColor = "E94560"

merge_title(ws6, 1, 1, 9,
    "TOP FREE RESOURCES IF STARTING FROM ZERO — RANKED BY IMPACT",
    BG_DARK, GOLD, 16, height=42)
merge_title(ws6, 2, 1, 9,
    "Total investment: ₹0 for courses. ~$50–100 in API credits. That's it. Less than one month of Netflix.",
    BG_DARK, GREEN_BRIGHT, 10, bold=False, height=24)

header_row(ws6, 3,
    ["Rank","Resource Name","Creator","Phase","Hours","Where to Find","Do ALL of it?","Why This Over Others","What You'll Build After"],
    [ACCENT_TEAL]*9, [TEAL_BRIGHT]*9, [10]*9, [True]*9, 32)

top_resources = [
    ("1","Neural Networks: Zero to Hero","Andrej Karpathy","Phase 3","14h",
     "youtube.com — search 'Karpathy Zero to Hero'","All 8 videos — no skips",
     "Best LLM intuition on the entire internet. Karpathy built GPT-2. He explains things no course teaches. This alone separates you from 90% of candidates who only did tutorials.",
     "Deep understanding of why attention works, how language models are trained, what context windows really mean"),
    ("2","LLM Course (GitHub roadmap)","Maxime Labonne / mlabonne","Phase 3","30h+",
     "github.com/mlabonne/llm-course","Follow the LLM Engineer path — skip Scientist path sections",
     "The most comprehensive free LLM learning roadmap on the internet. 200K+ GitHub stars. Includes notebooks, papers, and curated resources for every LLM engineering topic.",
     "Complete LLM engineering foundation: RAG, fine-tuning, deployment, evaluation — all with working notebooks"),
    ("3","Python for Everybody (freeCodeCamp)","Dr. Chuck / University of Michigan","Phase 1","16h",
     "youtube.com/freecodecamp — search 'Python for Everybody'","First 10 hours only (skip web frameworks)",
     "University-quality Python taught by one of the most engaging instructors online. Free. Clear. Modern. Better than most paid courses for total beginners.",
     "CLI tools, web scrapers, REST API integrations, CSV/JSON pipelines — the Python backbone for every AI project"),
    ("4","DeepLearning.AI Short Courses","Andrew Ng + OpenAI/Google/Anthropic","Phase 3","20h",
     "deeplearning.ai/short-courses — free audit","Do: Prompt Engineering, LangChain, RAG courses. Skip: basic intro courses",
     "Co-authored with OpenAI, Anthropic, Google — you're learning from the people who built the tools. These set the standard for LLM engineering education.",
     "Prompt templates, LangChain RAG pipeline, production-ready agent patterns — all from primary sources"),
    ("5","TechWorld with Nana — Docker Tutorial","Nana Janashia","Phase 5","4h",
     "youtube.com — search 'TechWorld Nana Docker Tutorial'","Watch full Docker tutorial; skip Swarm sections",
     "Nana is the best DevOps educator on YouTube. Clearer than most Udemy courses. Always updated. Used by 1M+ developers. Free.",
     "Dockerised FastAPI + LLM app you can actually deploy — the difference between a GitHub project and a live demo"),
    ("6","freeCodeCamp ML Course","Patrick Loeber","Phase 2","6h",
     "youtube.com/freecodecamp — search 'Machine Learning Python'","Python sections only (ignore any R content)",
     "Concise, practical ML with scikit-learn. Covers everything needed for LLM path context without time-wasting deep dives. StatQuest fills the conceptual gaps.",
     "Churn predictor, house price regressor, K-Means customer segmentation — the ML intuition you'll reference in interviews"),
    ("7","LangChain YouTube Channel","LangChain (Harrison Chase)","Phase 4A","8h",
     "youtube.com/c/LangChain","LangGraph tutorials + agent pattern videos",
     "Always more up-to-date than any recorded course. LangChain rewrites itself every 6 months — official channel keeps pace. The LangGraph series is the best agentic AI content available.",
     "Production multi-agent system using latest LangGraph patterns — more current than any Udemy course"),
    ("8","StatQuest with Josh Starmer","Josh Starmer","Phase 2–3","10h",
     "youtube.com — search 'StatQuest ML' + 'StatQuest attention'","Attention/transformer videos (required) + ML basics (as needed)",
     "Josh makes complex statistics and ML concepts genuinely understandable. His attention mechanism explanation is the best in existence. Watch when concepts from other resources confuse you.",
     "Intuitive understanding of attention, embeddings, and model evaluation — what lets you answer interview questions conceptually"),
    ("9","n8n Official YouTube Channel","n8n.io","Phase 4B","8h",
     "youtube.com/n8n","Workflow automation + AI node tutorials",
     "n8n covers their own tool better than any third-party course. Self-hosted = completely free. Growing adoption in companies. Differentiation on your CV from LangChain-only candidates.",
     "Business automation workflows that connect AI to real systems: email, databases, APIs, webhooks"),
    ("10","Hugging Face Course","Hugging Face","Phase 3","10h",
     "huggingface.co/learn/nlp-course","NLP Course chapters 1–4 + Fine-tuning chapter",
     "Free, authoritative, built by the people who maintain the transformers library. Better than any paid course for Hugging Face-specific knowledge. Certificate on completion (free).",
     "Running local models, building a fine-tuning pipeline, deploying on HF Spaces for free portfolio demos"),
]

for i, row in enumerate(top_resources, 4):
    rank, name, creator, phase, hours, location, do_all, why, builds = row
    bg = BG_PANEL if i % 2 == 0 else "111111"
    rank_bgs = ["B8860B","C0C0C0","CD7F32","1B4332","1B4332","1B4332","1B4332","1B4332","1B4332","1B4332"]
    rank_colors = [BG_DARK, BG_DARK, BG_DARK] + [LEARN_TEXT]*7
    rbg = rank_bgs[i-4] if i-4 < len(rank_bgs) else "1B4332"
    rfc = rank_colors[i-4] if i-4 < len(rank_colors) else LEARN_TEXT

    cells = [
        (rank,     True,  14, rfc,         rbg,         "center", False),
        (name,     True,  10, WHITE,        bg,          "left",   True),
        (creator,  False, 9,  TEAL_BRIGHT,  bg,          "left",   True),
        (phase,    False, 9,  ORANGE,       bg,          "center", False),
        (hours,    True,  10, GREEN_BRIGHT, FREE_TEAL,   "center", False),
        (location, False, 8,  YELLOW,       bg,          "left",   True),
        (do_all,   False, 8,  LIGHT_GRAY,   bg,          "left",   True),
        (why,      False, 8,  WHITE,        bg,          "left",   True),
        (builds,   False, 8,  ORANGE,       bg,          "left",   True),
    ]
    fill_row(ws6, i, cells, height=72)

set_col_widths(ws6, [6, 26, 22, 10, 8, 30, 22, 42, 40])

# Also: channels to avoid section
ws6.row_dimensions[15].height = 16
merge_title(ws6, 16, 1, 9,
    "CHANNELS / RESOURCES TO AVOID FOR THIS SPECIFIC PATH",
    SKIP_RED, SKIP_TEXT, 12, height=28)
header_row(ws6, 17,
    ["Resource","Creator","Why People Watch","Why to Skip","What to Do Instead","Exception"],
    [SKIP_RED]*6, [SKIP_TEXT]*6, [9]*6, [True]*6, 28)

avoid_data = [
    ("'ChatGPT/AI Tutorial' YouTube shorts (2–10 min)","Various viral creators","Quick AI tips, 'prompt hacks'","Surface-level content that doesn't teach engineering. Zero depth, maximum engagement bait.","Ed Donner LLM course or mlabonne roadmap for real engineering depth","Nobody — these don't teach skills"),
    ("Deep Learning Specialization (Coursera)","Andrew Ng","Prestigious, comprehensive deep learning","Teaches you to build neural networks from scratch. AI engineers use Hugging Face, not hand-rolled backprop. 60+ hours of content you won't use.","Karpathy Zero to Hero for LLM intuition (14h) instead of DL from scratch (60h+)","If you want to do ML research or PhD"),
    ("TensorFlow/Keras full courses","Various","TF is widely used in industry","LLM engineers use Hugging Face which runs on PyTorch. TensorFlow is declining in the LLM space. Wrong tool for this path.","Hugging Face transformers library in mlabonne course (free)","Computer vision or embedded ML roles"),
    ("'Passive income with AI' / 'Make money with AI' channels","Various","Side income motivation","These are marketing content, not engineering education. You'll learn to use tools, not build systems.","Any resource on this sheet — build real skills, not chatbot wrappers","Nobody"),
    ("Kaggle ML tutorials (for AI engineering)","Kaggle","Hands-on datasets, competitions","Kaggle skills ≠ AI engineering skills. Great for data science; wrong mental model for LLM/agent work.","mlabonne/llm-course notebooks for hands-on LLM practice","If you want data science / ML competition skills (different path)"),
    ("100+ hour 'complete AI bootcamp' courses","Udemy / Coursera various","Comprehensive coverage","The breadth means shallow depth everywhere. You'll finish 100 hours with no deep skills and no portfolio.","5-course focused path in this sheet OR 10 free resources above — depth over breadth","Nobody — pick focused resources"),
]

for i, row in enumerate(avoid_data, 18):
    resource, creator, why_watch, why_skip, instead, exception = row
    bg = BG_PANEL if i % 2 == 0 else "111111"
    cells = [
        (resource,   True,  9,  SKIP_TEXT,    SKIP_RED, "left",   True),
        (creator,    False, 9,  LIGHT_GRAY,   bg,       "left",   False),
        (why_watch,  False, 8,  MID_GRAY,     bg,       "left",   True),
        (why_skip,   False, 8,  SKIP_TEXT,    bg,       "left",   True),
        (instead,    False, 8,  GREEN_BRIGHT, bg,       "left",   True),
        (exception,  False, 8,  SKIM_TEXT,    SKIM_AMBER, "left", True),
    ]
    fill_row(ws6, i, cells, height=52)

set_col_widths(ws6, [6, 26, 22, 10, 8, 30, 22, 42, 40])
# Fix col widths for avoid section
for col, width in enumerate([28, 20, 22, 35, 30, 22], 1):
    ws6.column_dimensions[get_column_letter(col)].width = width

ws6.freeze_panes = "A4"

# ══════════════════════════════════════════════════════════════════════════════
# SHEET 7 — Mistakes to Avoid (same as Udemy, free-path annotated)
# ══════════════════════════════════════════════════════════════════════════════
ws7 = wb.create_sheet("⚠ Mistakes to Avoid")
ws7.sheet_view.showGridLines = False
ws7.sheet_properties.tabColor = "FF6B35"

merge_title(ws7, 1, 1, 8,
    "CRITICAL MISTAKES BEGINNERS MAKE ON THE FREE PATH — AND HOW TO AVOID THEM",
    BG_DARK, ORANGE, 16, height=42)
merge_title(ws7, 2, 1, 8,
    "Free path has one extra risk vs paid: no structure means more ways to go off track. These mistakes apply doubly.",
    BG_DARK, LIGHT_GRAY, 9, bold=False, height=24)

header_row(ws7, 3,
    ["#","Mistake","Category","How Common","What Beginners Do","What to Do Instead","Free-Path Specific Risk","Phase"],
    [ACCENT_TEAL]*8, [TEAL_BRIGHT]*8, [10]*8, [True]*8, 30)

mistakes_data = [
    ("1","Treating YouTube as background noise","Focus","Extremely Common on free path","Watch ML tutorials while multitasking. 'Learn' 50 hours without retaining anything.",
     "Code alongside every video. If you can't code what they're showing, pause and do it. Passive watching = zero skill.",
     "Free path has no accountability. Nobody checks if you're actually coding. Self-test: can you rebuild it from scratch after?","All phases"),
    ("2","Jumping between 10 resources without finishing one","Curriculum","Extremely Common","Start freeCodeCamp Python, switch to CS50P, switch to Mosh, switch to Corey Schafer. Finish nothing.",
     "Pick ONE primary resource per phase (see Sheet 6 rankings). Finish it before supplementing. Depth > variety.",
     "YouTube algorithm will surface new tutorials constantly. This is designed to keep you watching, not building. Resist.","Phase 1–2"),
    ("3","Skipping project builds because there are no deadlines","Portfolio","Very Common","Watch all the videos. Take notes. Never deploy anything. Have nothing on GitHub.",
     "Set your own deadlines. After every phase: BUILD A PROJECT before moving on. No exceptions. One GitHub push per week minimum.",
     "Paid courses have project checkpoints. Free courses don't. You must impose this discipline yourself.","All phases"),
    ("4","Using outdated YouTube tutorials","Relevance","Common on free path","Follow a 2022 LangChain tutorial. Learn deprecated APIs. Break everything. Get discouraged.",
     "Always check video upload date. For LangChain/LangGraph: only watch 2024+ content. Official channels are safest.",
     "YouTube doesn't expire old content. A 2021 LangChain tutorial is still in search results — and completely wrong for 2026.","Phase 3–4"),
    ("5","Course collecting without building (free version)","Portfolio","Extremely Common","Collect 20 free course certificates (Kaggle, DeepLearning.AI, HF, etc.) with zero projects.",
     "Max 2 certificates per phase. Projects beat certificates 100% of the time in LLM engineering interviews.",
     "Free courses are easy to start and get certificates. It feels like progress. It isn't. Build something.","Phase 1–3"),
    ("6","Ignoring documentation in favour of more YouTube","Skills","Very Common","When stuck, search YouTube for a tutorial instead of reading official docs.",
     "Every tool you use: read the Getting Started docs first. Docs are authoritative, YouTube is not. GitHub issues > YouTube for debugging.",
     "Documentation is free and more accurate than YouTube. Ability to read docs is a job skill interviewers test.","All phases"),
    ("7","Skipping evals and observability because the tools seem complicated","Production readiness","Very Common on free path","Build a working demo. Never add LangSmith tracing because it 'seems complicated'. Can't debug anything.",
     "Set up LangSmith (free tier) on Day 1 of Phase 4. It's 3 lines of code. Every project after Phase 3 must have tracing.",
     "LangSmith is free for 5K traces/month. There is no cost reason to skip this. Only reason is laziness.","Phase 3–4"),
    ("8","Never deploying because cloud setup seems hard","Portfolio","Very Common","Every project runs on localhost. Nothing is shareable. Can't demo in an interview.",
     "Deploy ONE project on Vercel or Railway free tier (5 min setup). One live URL is worth 10 GitHub repos.",
     "GCP gives $300 free credit. Railway/Vercel free tier requires no credit card. Zero cost barrier to deployment.","Phase 5"),
    ("9","Confusing consumption with learning","Mindset","Extremely Common","Watch 200 hours of YouTube. Feel like you've learned. Can't code anything without the tutorial open.",
     "Feynman technique: after each resource, close it and explain the concept to yourself. Then code it from memory.",
     "Free path has infinite content to consume. The dopamine hit of watching feels like progress. Test yourself constantly.","All phases"),
    ("10","Not tracking your own progress","Organisation","Common on free path","Rely on YouTube history as a 'curriculum'. No structured notes, no GitHub commits, no project log.",
     "Use a simple Notion or Obsidian daily log. Track: what you learned, what you built, what confused you. Visible progress = motivation.",
     "Paid courses have progress bars. Free path has nothing. Create your own system or you'll underestimate how far you've come.","All phases"),
    ("11","Applying before Phase 3 is complete","Job search timing","Common","Apply for AI engineer roles after finishing Phase 1–2. Get rejected. Get demoralised.",
     "Wait until Phase 3 is complete with at least 3 solid GitHub projects. Quality applications beat quantity.",
     "Free path tempts you to apply early because you feel like you're 'almost there'. You're not. Phase 3 is the minimum bar.","Phase 1–2"),
    ("12","Using free path as an excuse to avoid API costs","Cost","Common","Skip all API integration because 'I can't afford it'. Build demo chatbots that call no real APIs.",
     "Budget $30 for OpenAI credits. Use gpt-4o-mini (cheap) for dev. $30 is enough for 6 months of learning if used smartly.",
     "API costs are the one genuine cost on the free path. Don't avoid them — build cost awareness into your workflow from Day 1.","Phase 3"),
]

for i, row in enumerate(mistakes_data, 4):
    num, mistake, category, common, what, should, free_risk, phase = row
    bg = BG_PANEL if i % 2 == 0 else "111111"
    cells = [
        (num,      True,  11, ORANGE,      bg,       "center", False),
        (mistake,  True,  9,  WHITE,        bg,       "left",   True),
        (category, False, 8,  TEAL_BRIGHT,  bg,       "center", True),
        (common,   False, 8,  SKIP_TEXT,    SKIP_RED, "left",   True),
        (what,     False, 8,  LIGHT_GRAY,   bg,       "left",   True),
        (should,   False, 8,  GREEN_BRIGHT, LEARN_GREEN, "left", True),
        (free_risk,True,  8,  SKIM_TEXT,    SKIM_AMBER, "left", True),
        (phase,    True,  9,  TEAL_BRIGHT,  ACCENT_TEAL, "center", False),
    ]
    fill_row(ws7, i, cells, height=65)

set_col_widths(ws7, [4, 24, 15, 16, 35, 35, 35, 12])
ws7.freeze_panes = "A4"

# ── Save ──────────────────────────────────────────────────────────────────────
out = "AI_Engineer_Free_YouTube_Roadmap_2026.xlsx"
wb.save(out)
print("Saved:", out)